import requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone
import os
import statsmodels.api as sm
from statsmodels.tsa.arima.model import ARIMA
from sklearn.ensemble import GradientBoostingRegressor, RandomForestRegressor
import dash
from dash import dcc, html
from dash.dependencies import Input, Output
import plotly.express as px
import plotly.graph_objects as go
import logging
from concurrent.futures import ThreadPoolExecutor
import time
from pmdarima import auto_arima
from sklearn.model_selection import GridSearchCV
from plotly.subplots import make_subplots
import ccxt
from sklearn.linear_model import LinearRegression
from sklearn.svm import SVR
from sklearn.preprocessing import StandardScaler
import warnings
import scipy.stats as stats
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score
warnings.filterwarnings('ignore')

# Logging Configuration
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# API Configuration & Global Variables
API_KEY = os.getenv("COINAPI_KEY", "#####")
BASE_URL = "https://rest.coinapi.io/v1/"

global_data = {}       # Stores fetched OHLCV data for each asset
global_metrics = {}    # Stores calculated risk and predictive metrics
fetch_cache = {}       # Cache for API calls

# Global model configuration
MODEL_CONFIG = {
    "ARIMA_ORDER": (1, 0, 1),
    "RF_N_ESTIMATORS": 100,
    "RF_MAX_DEPTH": 10,
    "RF_MIN_SAMPLES_SPLIT": 5
}

# Data Fetching and Processing Functions
def fetch_ohlcv(symbol_id, period_id, time_start, time_end, retries=3):
    cache_key = (symbol_id, period_id, time_start, time_end)
    if cache_key in fetch_cache:
        logging.info("Fetch cache hit for %s", cache_key)
        return fetch_cache[cache_key]
    
    endpoint = f"{BASE_URL}ohlcv/{symbol_id}/history"
    headers = {"X-CoinAPI-Key": API_KEY}
    params = {
        "period_id": period_id,
        "time_start": time_start,
        "time_end": time_end,
        "limit": 10000,
        "include_empty_items": "true"
    }
    for attempt in range(retries):
        try:
            response = requests.get(endpoint, headers=headers, params=params)
            response.raise_for_status()
            df = pd.DataFrame(response.json())
            df['time_period_start'] = pd.to_datetime(df['time_period_start'])
            fetch_cache[cache_key] = df
            logging.info("Fetched data for %s and cached.", symbol_id)
            return df
        except Exception as e:
            if attempt < retries - 1:
                wait_time = 2 ** attempt
                logging.warning("API call failed, retrying in %d seconds...", wait_time)
                time.sleep(wait_time)
            else:
                logging.error("Error fetching data for %s after %d attempts: %s", symbol_id, retries, e)
                raise e

def fetch_all_data(symbols, period, time_start, time_end):
    with ThreadPoolExecutor() as executor:
        futures = {executor.submit(fetch_ohlcv, symbol_id, period, time_start, time_end): asset for asset, symbol_id in symbols.items()}
        results = {}
        for future in futures:
            try:
                results[futures[future]] = future.result()
            except Exception as e:
                logging.error("Error fetching data for %s: %s", futures[future], e)
        return results

def compute_frequency_returns(df, freq):
    df = df.copy()
    df['time'] = pd.to_datetime(df['time_period_start'])
    df.set_index('time', inplace=True)
    df['price_close'] = pd.to_numeric(df['price_close'])
    price_series = df['price_close'].resample(freq).last().dropna()
    returns = np.log(price_series / price_series.shift(1)).dropna()
    return returns

def resample_prices(df, freq):
    df = df.copy()
    df['time'] = pd.to_datetime(df['time_period_start'])
    df.set_index('time', inplace=True)
    df['price_close'] = pd.to_numeric(df['price_close'])
    return df['price_close'].resample(freq).last().dropna()

# Technical Indicator Functions
def calculate_sma(series, window=30):
    return series.rolling(window=window).mean()

def calculate_rsi(series, window=14):
    delta = series.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

def calculate_bollinger_bands(series, window=20, num_std=2):
    rolling_mean = series.rolling(window=window).mean()
    rolling_std = series.rolling(window=window).std()
    upper_band = rolling_mean + (rolling_std * num_std)
    lower_band = rolling_mean - (rolling_std * num_std)
    return upper_band, lower_band

def calculate_rolling_correlation(alt_returns, bench_returns, window=30):
    if len(alt_returns) < window or len(bench_returns) < window:
        logging.warning("Insufficient data for rolling correlation")
        return pd.Series(dtype=float)
    merged = pd.merge(
        alt_returns.rename("alt_return"),
        bench_returns.rename("bench_return"),
        left_index=True, right_index=True, how='inner'
    )
    if len(merged) < window:
        logging.warning("Insufficient merged data for rolling correlation")
        return pd.Series(dtype=float)
    return merged['alt_return'].rolling(window=window).corr(merged['bench_return'])

# Risk Metrics Functions
def calculate_var(returns, confidence_level=0.95, method='historical', holding_period=1):
    if returns.empty:
        return None
    if method == 'historical':
        var = returns.quantile(1 - confidence_level)
    elif method == 'parametric':
        z_score = stats.norm.ppf(1 - confidence_level)
        var = returns.mean() - z_score * returns.std()
    # Scale for holding period
    var = var * np.sqrt(holding_period)
    return var

def calculate_beta_for_returns(alt_returns, bench_returns):
    merged = pd.merge(
        alt_returns.rename("alt_return"),
        bench_returns.rename("bench_return"),
        left_index=True, right_index=True, how='inner'
    )
    if merged.empty:
        return None
    beta = np.cov(merged['alt_return'], merged['bench_return'])[0, 1] / np.var(merged['bench_return'])
    return beta

def calculate_monthly_sharpe_ratio(returns, risk_free_rate=0.04305):
    if returns.empty:
        return None
    periods_per_year = 12
    monthly_rf = risk_free_rate / periods_per_year
    mean_return = returns.mean() - monthly_rf
    std_return = returns.std()
    return mean_return / std_return if std_return != 0 else np.nan

def calculate_correlation(alt_returns, bench_returns):
    merged = pd.merge(
        alt_returns.rename("alt_return"),
        bench_returns.rename("bench_return"),
        left_index=True, right_index=True, how='inner'
    )
    return merged['alt_return'].corr(merged['bench_return']) if not merged.empty else None

def calculate_max_drawdown(price_series):
    if price_series.empty:
        return None
    running_max = price_series.cummax()
    drawdown = price_series / running_max - 1
    return drawdown.min()

def calculate_volatility(returns, periods_per_year=365):
    if returns.empty:
        return None
    return returns.std() * np.sqrt(periods_per_year)

def calculate_alpha(alt_returns, bench_returns, risk_free_rate=0):
    merged = pd.merge(
        alt_returns.rename("alt_return"),
        bench_returns.rename("bench_return"),
        left_index=True, right_index=True, how='inner'
    )
    if merged.empty:
        return None
    X = sm.add_constant(merged['bench_return'])
    model = sm.OLS(merged['alt_return'], X).fit()
    return model.params.iloc[0]

def calculate_rolling_beta(alt_returns, bench_returns, window=30):
    if len(alt_returns) < window or len(bench_returns) < window:
        logging.warning("Insufficient data for rolling beta")
        return pd.Series(dtype=float)
    merged = pd.merge(
        alt_returns.rename("alt_return"),
        bench_returns.rename("bench_return"),
        left_index=True, right_index=True, how='inner'
    )
    if len(merged) < window:
        logging.warning("Insufficient merged data for rolling beta")
        return pd.Series(dtype=float)
    roll_cov = merged['alt_return'].rolling(window=window).cov(merged['bench_return'])
    roll_var = merged['bench_return'].rolling(window=window).var()
    return roll_cov / roll_var

def calculate_rolling_alpha(alt_returns, bench_returns, window=30):
    if len(alt_returns) < window or len(bench_returns) < window:
        logging.warning("Insufficient data for rolling alpha")
        return pd.Series(dtype=float)
    merged = pd.merge(
        alt_returns.rename("alt_return"),
        bench_returns.rename("bench_return"),
        left_index=True, right_index=True, how='inner'
    )
    if len(merged) < window:
        logging.warning("Insufficient merged data for rolling alpha")
        return pd.Series(dtype=float)
    alphas = []
    dates = []
    for i in range(window, len(merged) + 1):
        window_df = merged.iloc[i - window:i]
        X = sm.add_constant(window_df['bench_return'])
        model = sm.OLS(window_df['alt_return'], X).fit()
        alphas.append(model.params['const'])
        dates.append(window_df.index[-1])
    return pd.Series(alphas, index=dates)

# Helper Function for Feature Engineering
def compute_features_for_asset(asset, forecast_horizon=1, window=30):
    asset_data = global_data[asset]
    asset_returns = compute_frequency_returns(asset_data.copy(), 'D')
    btc_data = global_data["BTC"]
    btc_returns = compute_frequency_returns(btc_data.copy(), 'D')
    merged = pd.merge(
        asset_returns.rename("asset_return"),
        btc_returns.rename("btc_return"),
        left_index=True, right_index=True, how='inner'
    )
    df = merged.copy()
    df["lag_return"] = df["asset_return"].shift(1)
    df["rolling_vol"] = df["asset_return"].rolling(window=window).std()
    df["rolling_beta"] = calculate_rolling_beta(asset_returns, btc_returns, window=window)
    df["rolling_alpha"] = calculate_rolling_alpha(asset_returns, btc_returns, window=window)
    asset_prices = resample_prices(asset_data.copy(), 'D')
    df["sma"] = calculate_sma(asset_prices, window=30).loc[df.index]
    df["rsi"] = calculate_rsi(asset_prices, window=14).loc[df.index]
    upper_band, lower_band = calculate_bollinger_bands(asset_prices, window=20, num_std=2)
    df["bb_upper"] = upper_band.loc[df.index]
    df["bb_lower"] = lower_band.loc[df.index]
    df["rolling_corr"] = calculate_rolling_correlation(asset_returns, btc_returns, window=30)
    df = df.dropna()
    if df.empty:
        return None, None
    df["target"] = df["asset_return"].shift(-forecast_horizon)
    df = df.dropna()
    features = df[["lag_return", "rolling_vol", "rolling_beta", "rolling_alpha", "sma", "rsi", "bb_upper", "bb_lower", "rolling_corr"]]
    target = df["target"]
    if features.empty or target.empty:
        return None, None
    return features, target

# Tuning Function
def tune_models(asset):
    asset_returns = compute_frequency_returns(global_data[asset].copy(), 'D')
    if len(asset_returns) < 30:
        logging.warning(f"Insufficient data for ARIMA tuning for {asset}")
        arima_order = MODEL_CONFIG["ARIMA_ORDER"]
    else:
        try:
            arima_model = auto_arima(asset_returns, start_p=0, start_q=0, max_p=3, max_q=3, d=0, seasonal=False, trace=False)
            arima_order = arima_model.order
        except Exception as e:
            logging.error(f"ARIMA tuning failed for {asset}: {e}")
            arima_order = MODEL_CONFIG["ARIMA_ORDER"]
    
    features, target = compute_features_for_asset(asset)
    if features is None or len(features) < 30:
        logging.warning(f"Insufficient data for RF tuning for {asset}")
        rf_params = {
            'n_estimators': MODEL_CONFIG["RF_N_ESTIMATORS"],
            'max_depth': MODEL_CONFIG["RF_MAX_DEPTH"],
            'min_samples_split': MODEL_CONFIG["RF_MIN_SAMPLES_SPLIT"]
        }
    else:
        param_grid = {
            'n_estimators': [50, 100],
            'max_depth': [5, 10],
            'min_samples_split': [2, 5]
        }
        rf = RandomForestRegressor(random_state=42)
        grid_search = GridSearchCV(rf, param_grid, cv=3, scoring='neg_mean_squared_error', n_jobs=-1)
        try:
            grid_search.fit(features, target)
            rf_params = grid_search.best_params_
        except Exception as e:
            logging.error(f"RF tuning failed for {asset}: {e}")
            rf_params = {
                'n_estimators': MODEL_CONFIG["RF_N_ESTIMATORS"],
                'max_depth': MODEL_CONFIG["RF_MAX_DEPTH"],
                'min_samples_split': MODEL_CONFIG["RF_MIN_SAMPLES_SPLIT"]
            }
    return arima_order, rf_params

# Predictive Modeling Functions
def predict_returns_for_asset(asset, arima_order=None, rf_params=None, forecast_horizon=1, window=30):
    if asset not in global_data:
        return None, None, None, None
    if asset == "BTC":
        return 0, 0, 1, 0
    asset_data = global_data[asset]
    asset_returns = compute_frequency_returns(asset_data.copy(), 'D')
    btc_data = global_data["BTC"]
    btc_returns = compute_frequency_returns(btc_data.copy(), 'D')
    
    arima_order = arima_order if arima_order is not None else MODEL_CONFIG["ARIMA_ORDER"]
    rf_params = rf_params if rf_params is not None else {
        'n_estimators': MODEL_CONFIG["RF_N_ESTIMATORS"],
        'max_depth': MODEL_CONFIG["RF_MAX_DEPTH"],
        'min_samples_split': MODEL_CONFIG["RF_MIN_SAMPLES_SPLIT"]
    }
    
    try:
        arima_model = ARIMA(asset_returns, order=arima_order)
        arima_result = arima_model.fit()
        arima_forecast = arima_result.forecast(steps=forecast_horizon)
        arima_pred = arima_forecast.iloc[0] if not arima_forecast.empty else 0
    except Exception as e:
        logging.error("ARIMA error for %s: %s", asset, e)
        arima_pred = 0

    features, target = compute_features_for_asset(asset, forecast_horizon, window)
    if features is None:
        return None, None, None, None
    
    X_train, X_test, y_train, y_test = train_test_split(
        features,
        target,
        test_size=0.2,
        random_state=42
    )
    
    rf_model = RandomForestRegressor(**rf_params, random_state=42)
    rf_model.fit(X_train, y_train)
    rf_pred = rf_model.predict(X_test)
    rf_rmse = np.sqrt(mean_squared_error(y_test, rf_pred))
    rf_r2 = r2_score(y_test, rf_pred)
    rf_latest_pred = rf_model.predict(features.iloc[-1].values.reshape(1, -1))[0]
    
    ensemble_pred = (arima_pred + rf_latest_pred) / 2
    var_95 = calculate_var(asset_returns)
    return ensemble_pred, rf_rmse, rf_r2, var_95

def predict_linear_regression(
    asset: str,
    forecast_horizon: int = 1,
    window: int = 30
) -> tuple:
    """
    Predict returns using Linear Regression.
    
    Args:
        asset: Asset symbol
        forecast_horizon: Days ahead to forecast
        window: Rolling window size
    
    Returns:
        tuple: (prediction, rmse, r2, var_95)
    """
    if asset not in global_data:
        return None, None, None, None
        
    features, target = compute_features_for_asset(asset, forecast_horizon, window)
    if features is None:
        return None, None, None, None
        
    X_train, X_test, y_train, y_test = train_test_split(
        features,
        target,
        test_size=0.2,
        random_state=42
    )
    
    model = LinearRegression()
    model.fit(X_train, y_train)
    predictions = model.predict(X_test)
    rmse = np.sqrt(mean_squared_error(y_test, predictions))
    r2 = r2_score(y_test, predictions)
    latest_pred = model.predict(features.iloc[-1].values.reshape(1, -1))[0]
    
    asset_returns = compute_frequency_returns(global_data[asset].copy(), 'D')
    var_95 = calculate_var(asset_returns)
    
    return latest_pred, rmse, r2, var_95

def predict_svr(
    asset: str,
    forecast_horizon: int = 1,
    window: int = 30
) -> tuple:
    """
    Predict returns using Support Vector Regression.
    
    Args:
        asset: Asset symbol
        forecast_horizon: Days ahead to forecast
        window: Rolling window size
    
    Returns:
        tuple: (prediction, rmse, r2, var_95)
    """
    if asset not in global_data:
        return None, None, None, None
        
    features, target = compute_features_for_asset(asset, forecast_horizon, window)
    if features is None:
        return None, None, None, None
        
    X_train, X_test, y_train, y_test = train_test_split(
        features,
        target,
        test_size=0.2,
        random_state=42
    )
    
    model = SVR()
    model.fit(X_train, y_train)
    predictions = model.predict(X_test)
    rmse = np.sqrt(mean_squared_error(y_test, predictions))
    r2 = r2_score(y_test, predictions)
    latest_pred = model.predict(features.iloc[-1].values.reshape(1, -1))[0]
    
    asset_returns = compute_frequency_returns(global_data[asset].copy(), 'D')
    var_95 = calculate_var(asset_returns)
    
    return latest_pred, rmse, r2, var_95

def predict_gbm(asset, forecast_horizon=1, window=30):
    if asset not in global_data:
        return None, None, None, None
    features, target = compute_features_for_asset(asset, forecast_horizon, window)
    if features is None:
        return None, None, None, None
    from sklearn.model_selection import train_test_split
    from sklearn.metrics import mean_squared_error, r2_score
    X_train, X_test, y_train, y_test = train_test_split(features, target, test_size=0.2, random_state=42)
    model = GradientBoostingRegressor(random_state=42)
    model.fit(X_train, y_train)
    predictions = model.predict(X_test)
    rmse = np.sqrt(mean_squared_error(y_test, predictions))
    r2 = r2_score(y_test, predictions)
    latest_pred = model.predict(features.iloc[-1].values.reshape(1, -1))[0]
    asset_returns = compute_frequency_returns(global_data[asset].copy(), 'D')
    var_95 = calculate_var(asset_returns)
    return latest_pred, rmse, r2, var_95

# Data Update Function
def update_data(days=90, period="1DAY"):
    global global_data, global_metrics
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=days)
    time_start = start_date.strftime("%Y-%m-%dT%H:%M:%SZ")
    time_end = end_date.strftime("%Y-%m-%dT%H:%M:%SZ")

    symbols = {
        "BTC": "BINANCE_SPOT_BTC_USDT",
        "ETH": "BINANCE_SPOT_ETH_USDT",
        "XRP": "BINANCE_SPOT_XRP_USDT",
        "ADA": "BINANCE_SPOT_ADA_USDT",
        "SOL": "BINANCE_SPOT_SOL_USDT",
        "HBAR": "BINANCE_SPOT_HBAR_USDT",
        "CRV": "BINANCE_SPOT_CRV_USDT",
        "DOGE": "BINANCE_SPOT_DOGE_USDT",
        "SUI": "BINANCE_SPOT_SUI_USDT",
        "TON": "BINANCE_SPOT_TON_USDT",
        "SHIB": "BINANCE_SPOT_SHIB_USDT",
        "AAVE": "BINANCE_SPOT_AAVE_USDT"
    }

    logging.info("Fetching data for all assets with period %s and days %d...", period, days)
    global_data = fetch_all_data(symbols, period, time_start, time_end)
    logging.info("Data fetched for all assets.")

    btc_data = global_data["BTC"]
    btc_daily_returns = compute_frequency_returns(btc_data.copy(), 'D')
    btc_weekly_returns = compute_frequency_returns(btc_data.copy(), 'W')
    btc_monthly_returns = compute_frequency_returns(btc_data.copy(), 'ME')

    metrics = {}
    for asset in symbols.keys():
        if asset not in global_data:
            continue
        alt_data = global_data[asset]
        alt_daily_returns = compute_frequency_returns(alt_data.copy(), 'D')
        alt_weekly_returns = compute_frequency_returns(alt_data.copy(), 'W')
        alt_monthly_returns = compute_frequency_returns(alt_data.copy(), 'ME')

        daily_roll_beta_series = calculate_rolling_beta(alt_daily_returns, btc_daily_returns, window=30)
        daily_roll_beta = daily_roll_beta_series.iloc[-1] if not daily_roll_beta_series.empty else None
        weekly_roll_beta_series = calculate_rolling_beta(alt_weekly_returns, btc_weekly_returns, window=4)
        weekly_roll_beta = weekly_roll_beta_series.iloc[-1] if not weekly_roll_beta_series.empty else None
        monthly_roll_beta_series = calculate_rolling_beta(alt_monthly_returns, btc_monthly_returns, window=3)
        monthly_roll_beta = monthly_roll_beta_series.iloc[-1] if not monthly_roll_beta_series.empty else None
        daily_roll_alpha_series = calculate_rolling_alpha(alt_daily_returns, btc_daily_returns, window=30)
        daily_roll_alpha = daily_roll_alpha_series.iloc[-1] if not daily_roll_alpha_series.empty else None
        daily_roll_corr_series = calculate_rolling_correlation(alt_daily_returns, btc_daily_returns, window=30)
        daily_roll_corr = daily_roll_corr_series.iloc[-1] if not daily_roll_corr_series.empty else None
        daily_roll_vol_series = alt_daily_returns.rolling(window=30).std()
        daily_roll_vol = daily_roll_vol_series.iloc[-1] if not daily_roll_vol_series.empty else None
        weekly_roll_alpha_series = calculate_rolling_alpha(alt_weekly_returns, btc_weekly_returns, window=4)
        weekly_roll_alpha = weekly_roll_alpha_series.iloc[-1] if not weekly_roll_alpha_series.empty else None
        weekly_roll_corr_series = calculate_rolling_correlation(alt_weekly_returns, btc_weekly_returns, window=4)
        weekly_roll_corr = weekly_roll_corr_series.iloc[-1] if not weekly_roll_corr_series.empty else None
        weekly_roll_vol_series = alt_weekly_returns.rolling(window=4).std()
        weekly_roll_vol = weekly_roll_vol_series.iloc[-1] if not weekly_roll_vol_series.empty else None
        monthly_roll_alpha_series = calculate_rolling_alpha(alt_monthly_returns, btc_monthly_returns, window=3)
        monthly_roll_alpha = monthly_roll_alpha_series.iloc[-1] if not monthly_roll_alpha_series.empty else None
        monthly_roll_corr_series = calculate_rolling_correlation(alt_monthly_returns, btc_monthly_returns, window=3)
        monthly_roll_corr = monthly_roll_corr_series.iloc[-1] if not monthly_roll_corr_series.empty else None
        monthly_roll_vol_series = alt_monthly_returns.rolling(window=3).std()
        monthly_roll_vol = monthly_roll_vol_series.iloc[-1] if not monthly_roll_vol_series.empty else None

        metrics[asset] = {
            "Daily": {
                "Beta": calculate_beta_for_returns(alt_daily_returns, btc_daily_returns),
                "Corr": calculate_correlation(alt_daily_returns, btc_daily_returns),
                "Max Drawdown": calculate_max_drawdown(resample_prices(alt_data.copy(), 'D')),
                "Volatility": calculate_volatility(alt_daily_returns),
                "Alpha": calculate_alpha(alt_daily_returns, btc_daily_returns),
                "Rolling Beta": daily_roll_beta,
                "Rolling Alpha": daily_roll_alpha,
                "Rolling Corr": daily_roll_corr,
                "Rolling Volatility": daily_roll_vol,
            },
            "Weekly": {
                "Beta": calculate_beta_for_returns(alt_weekly_returns, btc_weekly_returns),
                "Corr": calculate_correlation(alt_weekly_returns, btc_weekly_returns),
                "Max Drawdown": calculate_max_drawdown(resample_prices(alt_data.copy(), 'W')),
                "Volatility": calculate_volatility(alt_weekly_returns),
                "Alpha": calculate_alpha(alt_weekly_returns, btc_weekly_returns),
                "Rolling Beta": weekly_roll_beta,
                "Rolling Alpha": weekly_roll_alpha,
                "Rolling Corr": weekly_roll_corr,
                "Rolling Volatility": weekly_roll_vol,
            },
            "Monthly": {
                "Beta": calculate_beta_for_returns(alt_monthly_returns, btc_monthly_returns),
                "Corr": calculate_correlation(alt_monthly_returns, btc_monthly_returns),
                "Max Drawdown": calculate_max_drawdown(resample_prices(alt_data.copy(), 'ME')),
                "Sharpe": calculate_monthly_sharpe_ratio(alt_monthly_returns),
                "Volatility": calculate_volatility(alt_monthly_returns),
                "Alpha": calculate_alpha(alt_monthly_returns, btc_monthly_returns),
                "Rolling Beta": monthly_roll_beta,
                "Rolling Alpha": monthly_roll_alpha,
                "Rolling Corr": monthly_roll_corr,
                "Rolling Volatility": monthly_roll_vol,
            }
        }

        arima_order, rf_params = tune_models(asset)
        metrics[asset]["arima_order"] = arima_order
        metrics[asset]["rf_params"] = rf_params

    for asset in metrics.keys():
        metrics[asset]["Current Price"] = global_data[asset]['price_close'].iloc[-1]
        pred_return, pred_rmse, pred_r2, var_95 = predict_returns_for_asset(asset, metrics[asset]["arima_order"], metrics[asset]["rf_params"])
        lin_pred, lin_rmse, lin_r2, lin_var = predict_linear_regression(asset)
        svr_pred, svr_rmse, svr_r2, svr_var = predict_svr(asset)
        gbm_pred, gbm_rmse, gbm_r2, gbm_var = predict_gbm(asset)
        metrics[asset]["Predicted"] = {
            "Predicted Return": pred_return,
            "Predicted RMSE": pred_rmse,
            "Predicted R2": pred_r2,
            "VaR (95%)": var_95
        }
        metrics[asset]["Additional"] = {
            "GBM Predicted Return": gbm_pred,
            "GBM RMSE": gbm_rmse,
            "GBM R2": gbm_r2,
            "GBM VaR": gbm_var,
            "Lin Predicted Return": lin_pred,
            "Lin RMSE": lin_rmse,
            "Lin R2": lin_r2,
            "Lin VaR": lin_var,
            "SVR Predicted Return": svr_pred,
            "SVR RMSE": svr_rmse,
            "SVR R2": svr_r2,
            "SVR VaR": svr_var
        }
    
    global_metrics = metrics
    run_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    dashboard_df = get_dashboard_df(metrics)
    append_dashboard_to_excel(dashboard_df, run_timestamp)
    logging.info("Data updated at: %s", run_timestamp)
    print("Data updated at:", run_timestamp)

# Excel Export Functions
def get_dashboard_df(metrics_dict):
    metric_columns = [
        "Daily Beta", "Weekly Beta", "Monthly Beta",
        "Monthly Sharpe",
        "Daily Corr", "Weekly Corr", "Monthly Corr",
        "Daily Drawdown", "Weekly Drawdown", "Monthly Drawdown",
        "Daily Volatility", "Weekly Volatility", "Monthly Volatility",
        "Daily Alpha", "Weekly Alpha", "Monthly Alpha",
        "Daily Rolling Beta", "Weekly Rolling Beta", "Monthly Rolling Beta",
        "Daily Rolling Alpha", "Weekly Rolling Alpha", "Monthly Rolling Alpha",
        "Daily Rolling Corr", "Weekly Rolling Corr", "Monthly Rolling Corr",
        "Daily Rolling Volatility", "Weekly Rolling Volatility", "Monthly Rolling Volatility",
        "ARIMA + RF Predicted Return", "ARIMA + RF Predicted RMSE", "ARIMA + RF Predicted R2", "ARIMA + RF VaR (95%)",
        "ARIMA + RF Predicted Next Day Price",
        "GBM Predicted Return", "GBM Predicted RMSE", "GBM Predicted R2", "GBM VaR (95%)",
        "GBM Predicted Next Day Price",
        "Lin Predicted Return", "Lin Predicted RMSE", "Lin Predicted R2", "Lin VaR (95%)",
        "Lin Predicted Next Day Price",
        "SVR Predicted Return", "SVR Predicted RMSE", "SVR Predicted R2", "SVR VaR (95%)",
        "SVR Predicted Next Day Price",
        "Current Price",
        "ARIMA_p", "ARIMA_d", "ARIMA_q",
        "RF_n_estimators", "RF_max_depth", "RF_min_samples_split"
    ]
    assets = sorted([asset for asset in metrics_dict if asset != "BTC"])
    if "BTC" in metrics_dict:
        assets.append("BTC")
    data = {}
    for asset in assets:
        freq_metrics = metrics_dict.get(asset, {})
        current_price = freq_metrics.get("Current Price")
        row = {
            "Daily Beta": freq_metrics.get("Daily", {}).get("Beta"),
            "Weekly Beta": freq_metrics.get("Weekly", {}).get("Beta"),
            "Monthly Beta": freq_metrics.get("Monthly", {}).get("Beta"),
            "Monthly Sharpe": freq_metrics.get("Monthly", {}).get("Sharpe"),
            "Daily Corr": freq_metrics.get("Daily", {}).get("Corr"),
            "Weekly Corr": freq_metrics.get("Weekly", {}).get("Corr"),
            "Monthly Corr": freq_metrics.get("Monthly", {}).get("Corr"),
            "Daily Drawdown": freq_metrics.get("Daily", {}).get("Max Drawdown"),
            "Weekly Drawdown": freq_metrics.get("Weekly", {}).get("Max Drawdown"),
            "Monthly Drawdown": freq_metrics.get("Monthly", {}).get("Max Drawdown"),
            "Daily Volatility": freq_metrics.get("Daily", {}).get("Volatility"),
            "Weekly Volatility": freq_metrics.get("Weekly", {}).get("Volatility"),
            "Monthly Volatility": freq_metrics.get("Monthly", {}).get("Volatility"),
            "Daily Alpha": freq_metrics.get("Daily", {}).get("Alpha"),
            "Weekly Alpha": freq_metrics.get("Weekly", {}).get("Alpha"),
            "Monthly Alpha": freq_metrics.get("Monthly", {}).get("Alpha"),
            "Daily Rolling Beta": freq_metrics.get("Daily", {}).get("Rolling Beta"),
            "Weekly Rolling Beta": freq_metrics.get("Weekly", {}).get("Rolling Beta"),
            "Monthly Rolling Beta": freq_metrics.get("Monthly", {}).get("Rolling Beta"),
            "Daily Rolling Alpha": freq_metrics.get("Daily", {}).get("Rolling Alpha"),
            "Weekly Rolling Alpha": freq_metrics.get("Weekly", {}).get("Rolling Alpha"),
            "Monthly Rolling Alpha": freq_metrics.get("Monthly", {}).get("Rolling Alpha"),
            "Daily Rolling Corr": freq_metrics.get("Daily", {}).get("Rolling Corr"),
            "Weekly Rolling Corr": freq_metrics.get("Weekly", {}).get("Rolling Corr"),
            "Monthly Rolling Corr": freq_metrics.get("Monthly", {}).get("Rolling Corr"),
            "Daily Rolling Volatility": freq_metrics.get("Daily", {}).get("Rolling Volatility"),
            "Weekly Rolling Volatility": freq_metrics.get("Weekly", {}).get("Rolling Volatility"),
            "Monthly Rolling Volatility": freq_metrics.get("Monthly", {}).get("Rolling Volatility"),
            "ARIMA + RF Predicted Return": freq_metrics.get("Predicted", {}).get("Predicted Return"),
            "ARIMA + RF Predicted RMSE": freq_metrics.get("Predicted", {}).get("Predicted RMSE"),
            "ARIMA + RF Predicted R2": freq_metrics.get("Predicted", {}).get("Predicted R2"),
            "ARIMA + RF VaR (95%)": freq_metrics.get("Predicted", {}).get("VaR (95%)"),
            "GBM Predicted Return": freq_metrics.get("Additional", {}).get("GBM Predicted Return"),
            "GBM Predicted RMSE": freq_metrics.get("Additional", {}).get("GBM RMSE"),
            "GBM Predicted R2": freq_metrics.get("Additional", {}).get("GBM R2"),
            "GBM VaR (95%)": freq_metrics.get("Additional", {}).get("GBM VaR"),
            "Lin Predicted Return": freq_metrics.get("Additional", {}).get("Lin Predicted Return"),
            "Lin Predicted RMSE": freq_metrics.get("Additional", {}).get("Lin RMSE"),
            "Lin Predicted R2": freq_metrics.get("Additional", {}).get("Lin R2"),
            "Lin VaR (95%)": freq_metrics.get("Additional", {}).get("Lin VaR"),
            "SVR Predicted Return": freq_metrics.get("Additional", {}).get("SVR Predicted Return"),
            "SVR Predicted RMSE": freq_metrics.get("Additional", {}).get("SVR RMSE"),
            "SVR Predicted R2": freq_metrics.get("Additional", {}).get("SVR R2"),
            "SVR VaR (95%)": freq_metrics.get("Additional", {}).get("SVR VaR"),
            "Current Price": current_price,
            "ARIMA_p": freq_metrics.get("arima_order", (None, None, None))[0],
            "ARIMA_d": freq_metrics.get("arima_order", (None, None, None))[1],
            "ARIMA_q": freq_metrics.get("arima_order", (None, None, None))[2],
            "RF_n_estimators": freq_metrics.get("rf_params", {}).get("n_estimators"),
            "RF_max_depth": freq_metrics.get("rf_params", {}).get("max_depth"),
            "RF_min_samples_split": freq_metrics.get("rf_params", {}).get("min_samples_split")
        }
        if row["ARIMA + RF Predicted Return"] is not None and current_price is not None:
            row["ARIMA + RF Predicted Next Day Price"] = current_price * np.exp(row["ARIMA + RF Predicted Return"])
        else:
            row["ARIMA + RF Predicted Next Day Price"] = None
        if row["GBM Predicted Return"] is not None and current_price is not None:
            row["GBM Predicted Next Day Price"] = current_price * np.exp(row["GBM Predicted Return"])
        else:
            row["GBM Predicted Next Day Price"] = None
        if row["Lin Predicted Return"] is not None and current_price is not None:
            row["Lin Predicted Next Day Price"] = current_price * np.exp(row["Lin Predicted Return"])
        else:
            row["Lin Predicted Next Day Price"] = None
        if row["SVR Predicted Return"] is not None and current_price is not None:
            row["SVR Predicted Next Day Price"] = current_price * np.exp(row["SVR Predicted Return"])
        else:
            row["SVR Predicted Next Day Price"] = None
        data[asset] = row
    df = pd.DataFrame.from_dict(data, orient="index")
    return df[metric_columns]

def append_dashboard_to_excel(df_dashboard, run_timestamp, filename="dashboard_metrics.xlsx", mode="append"):
    header_df = pd.DataFrame([{"Asset": f"Run Timestamp: {run_timestamp}", 
                                **{col: "" for col in df_dashboard.columns}}])
    df_with_assets = df_dashboard.reset_index().rename(columns={'index': 'Asset'})
    section_header = pd.DataFrame([{"Asset": "Assets →", 
                                     **{col: col for col in df_dashboard.columns}}])
    metric_definitions = pd.DataFrame({
        "Asset": [
            "Metric Definitions:",
            "Beta: Volatility measure relative to BTC",
            "Sharpe: Risk-adjusted return metric",
            "Correlation: Price correlation with BTC",
            "Drawdown: Maximum decline from peak",
            "Volatility: Standard deviation of log returns",
            "Alpha: Regression intercept of returns vs. BTC",
            "Rolling metrics: Computed over a moving window",
            "ARIMA + RF Predicted Return: Ensemble (ARIMA + RF) predicted log return",
            "GBM Predicted Return: GBM predicted log return",
            "Lin Predicted Return: Linear Regression predicted log return",
            "SVR Predicted Return: SVR predicted log return",
            "Predicted Next Day Price: Forecasted price for each model",
            "Current Price: Most recent price"
        ]
    })
    for col in df_dashboard.columns:
        metric_definitions[col] = ""
    final_columns = ["Asset"] + list(df_dashboard.columns)
    metric_definitions = metric_definitions[final_columns]
    blank_row = pd.DataFrame([{col: "" for col in final_columns}])
    final_df = pd.concat([
        header_df,
        blank_row,
        section_header,
        df_with_assets,
        blank_row,
        metric_definitions
    ], ignore_index=True)
    try:
        if mode == "overwrite" or not os.path.exists(filename):
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
        else:
            from openpyxl import load_workbook
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                workbook = writer.book
                last_row = workbook['Sheet1'].max_row if 'Sheet1' in workbook.sheetnames else 0
                final_df.to_excel(writer, sheet_name='Sheet1', startrow=last_row + 2, index=False, header=False)
        logging.info("Dashboard successfully appended to %s", filename)
    except Exception as e:
        logging.error("Error appending/creating Excel file: %s", e)
        print(f"\nError appending/creating Excel file: {e}")

# Dash Dashboard Setup
app = dash.Dash(__name__)

app.layout = html.Div(id='main-div', children=[
    dcc.Checklist(
        id='theme-switch',
        options=[{'label': 'Night Mode', 'value': 'night'}],
        value=[],
        inline=True,
        style={'marginBottom': '10px'}
    ),
    html.H1("Altcoin Risk & Returns Monitoring Dashboard"),
    html.Div([
        dcc.Dropdown(
            id='asset-dropdown',
            options=[{'label': asset, 'value': asset} for asset in 
                     ["ETH", "XRP", "ADA", "SOL", "HBAR", "CRV", "DOGE", "SUI", "TON", "SHIB", "AAVE", "BTC"]],
            value='ETH',
            multi=False,
            style={'width': '200px', 'display': 'inline-block'}
        ),
        html.Button("Refresh Data", id="refresh-button", n_clicks=0, style={'marginLeft': '20px'}),
        html.Label("Historical Days (Default: 90 days):", 
                  id='historical-days-label',
                  style={'marginLeft': '20px'}),
        dcc.Dropdown(
            id='days-dropdown',
            options=[{'label': str(d), 'value': d} for d in [30, 60, 90, 180]],
            value=90,
            style={'width': '100px', 'display': 'inline-block', 'marginLeft': '10px'}
        ),
        html.Label("Data Frequency (Default: 1 day):", 
                  id='frequency-label',
                  style={'marginLeft': '20px'}),
        dcc.Dropdown(
            id='period-dropdown',
            options=[{'label': p, 'value': p} for p in ["1MIN", "1HRS", "1DAY"]],
            value="1DAY",
            style={'width': '100px', 'display': 'inline-block', 'marginLeft': '10px'}
        ),
        html.Div(id="last-update", style={'marginLeft': '20px', 'fontWeight': 'bold', 'display': 'inline-block'}),
        html.Div(id='current-price-header', style={'marginTop': '20px', 'fontSize': '24px', 'fontWeight': 'bold'}),
        html.Div(id='ensemble-models-container', children=[
            html.Div(id='arima-rf-metrics'),
            html.Div(id='gbm-metrics')
        ], style={'marginTop': '20px', 'padding': '10px'}),
        html.Div(id='additional-models-container', children=[
            html.Div(id='additional-predicted-metrics')
        ], style={'marginTop': '20px', 'padding': '10px'}),
        html.Div(id='technical-analysis-container', children=[
            html.H3("Technical Analysis", style={'marginBottom': '20px'}),
            dcc.Loading(children=[
                html.Div([
                    dcc.Graph(id='candlestick-volume-chart'),
                    dcc.Graph(id='technical-indicators-chart'),
                    dcc.Graph(id='bollinger-bands-chart')
                ])
            ], type="default")
        ], style={'marginTop': '30px', 'padding': '10px'}),
        
        html.Div(id='risk-metrics-container', children=[
            html.H3("Risk Metrics", style={'marginBottom': '20px'}),
            html.Div(id='risk-metrics-content')
        ], style={'marginTop': '30px', 'padding': '10px'}),
    ], style={'marginBottom': '20px'}),
    
    dcc.Loading(children=[dcc.Graph(id='rolling-beta-chart')], type="default"),
    dcc.Loading(children=[dcc.Graph(id='volatility-chart')], type="default"),
    dcc.Loading(children=[dcc.Graph(id='rolling-alpha-chart')], type="default"),
    dcc.Loading(children=[dcc.Graph(id='rolling-correlation-chart')], type="default"),
    dcc.Loading(children=[dcc.Graph(id='correlation-heatmap')], type="default"),
    
    html.Div(id='rolling-metrics-definitions', children=[
        html.H3("Rolling Metrics Definitions"),
        html.Ul([
            html.Li("Rolling Beta: Measures sensitivity of asset returns to BTC over a moving window."),
            html.Li("Rolling Volatility: Standard deviation of asset's returns over a moving window."),
            html.Li("Rolling Alpha: Excess return relative to BTC computed via regression over a moving window."),
            html.Li("Rolling Correlation: Correlation between asset and BTC returns over a moving window.")
        ])
    ], style={'marginTop': '30px', 'padding': '10px'}),
    
    html.Div(id='model-definitions', children=[
        html.H3("Model Definitions"),
        html.Ul([
            html.Li("ARIMA + Random Forest Ensemble: Uses historical returns via ARIMA and a set of engineered features via Random Forest to forecast next-day log returns."),
            html.Li("Gradient Boosting Machine (GBM): An ensemble method that builds decision trees sequentially to minimize prediction error."),
            html.Li("Linear Regression: Fits a linear model on engineered features to predict next-day log returns."),
            html.Li("Support Vector Regression (SVR): Utilizes support vector machines with non-linear kernels to capture complex patterns.")
        ]),
        html.P("Computed features include: Lag Return, Rolling Volatility, Rolling Beta, Rolling Alpha, SMA, RSI, Bollinger Bands, and Rolling Correlation.")
    ], style={'marginTop': '30px', 'padding': '10px'}),
    
    html.Div(id='performance-metrics-definitions', children=[
        html.H3("Performance Metrics Definitions"),
        html.Ul([
            html.Li("Value at Risk (VaR): Estimates the maximum potential loss over a period at a given confidence level (e.g., 95%)."),
            html.Li("Root Mean Squared Error (RMSE): Measures the average magnitude of prediction errors; lower is better."),
            html.Li("R²: Indicates the proportion of variance in the observed data explained by the model; closer to 1 is better.")
        ])
    ], style={'marginTop': '30px', 'padding': '10px'})
])

# Callbacks
@app.callback(
    [Output('asset-dropdown', 'style'),
     Output('refresh-button', 'style'),
     Output('days-dropdown', 'style'),
     Output('period-dropdown', 'style'),
     Output('historical-days-label', 'style'),
     Output('frequency-label', 'style')],
    Input('theme-switch', 'value')
)
def update_input_styles(theme_value):
    if 'night' in theme_value:
        dropdown_style = {
            'width': '200px',
            'display': 'inline-block',
            'backgroundColor': '#2b2b2b',
            'color': 'white',
            'border': '1px solid #666',
            'borderRadius': '6px',
            'transition': 'all 0.3s ease-in-out',
            'boxShadow': '0 2px 4px rgba(0,0,0,0.2)'
        }
        button_style = {
            'marginLeft': '20px',
            'backgroundColor': '#2b2b2b',
            'color': 'white',
            'border': '1px solid #666',
            'borderRadius': '6px',
            'padding': '8px 16px',
            'cursor': 'pointer',
            'transition': 'all 0.3s ease-in-out',
            'boxShadow': '0 2px 4px rgba(0,0,0,0.2)',
            ':hover': {
                'backgroundColor': '#3b3b3b',
                'transform': 'translateY(-1px)',
                'boxShadow': '0 4px 8px rgba(0,0,0,0.3)'
            }
        }
        label_style = {
            'marginLeft': '20px',
            'color': 'white',
            'transition': 'all 0.3s ease-in-out',
            'fontSize': '1rem',
            'fontWeight': '500'
        }
    else:
        dropdown_style = {
            'width': '200px',
            'display': 'inline-block',
            'backgroundColor': 'white',
            'color': '#2d3436',
            'border': '1px solid #dfe6e9',
            'borderRadius': '6px',
            'transition': 'all 0.3s ease-in-out',
            'boxShadow': '0 2px 4px rgba(0,0,0,0.1)'
        }
        button_style = {
            'marginLeft': '20px',
            'backgroundColor': 'white',
            'color': '#2d3436',
            'border': '1px solid #dfe6e9',
            'borderRadius': '6px',
            'padding': '8px 16px',
            'cursor': 'pointer',
            'transition': 'all 0.3s ease-in-out',
            'boxShadow': '0 2px 4px rgba(0,0,0,0.1)',
            ':hover': {
                'backgroundColor': '#f5f6fa',
                'transform': 'translateY(-1px)',
                'boxShadow': '0 4px 8px rgba(0,0,0,0.15)'
            }
        }
        label_style = {
            'marginLeft': '20px',
            'color': '#2d3436',
            'transition': 'all 0.3s ease-in-out',
            'fontSize': '1rem',
            'fontWeight': '500'
        }
    
    return dropdown_style, button_style, dropdown_style, dropdown_style, label_style, label_style

@app.callback(
    [Output('ensemble-models-container', 'style'),
     Output('additional-models-container', 'style'),
     Output('rolling-metrics-definitions', 'style'),
     Output('model-definitions', 'style'),
     Output('performance-metrics-definitions', 'style')],
    Input('theme-switch', 'value')
)
def update_container_styles(theme_value):
    base_style = {
        'marginTop': '20px',
        'padding': '20px',
        'borderRadius': '8px',
        'boxShadow': '0 2px 4px rgba(0,0,0,0.1)'
    }
    if 'night' in theme_value:
        container_style = {
            **base_style,
            'backgroundColor': '#1e1e1e',
            'border': '1px solid #444',
            'boxShadow': '0 2px 4px rgba(0,0,0,0.2)'
        }
    else:
        container_style = {
            **base_style,
            'backgroundColor': 'white',
            'border': '1px solid #ddd'
        }
    return container_style, container_style, container_style, container_style, container_style

@app.callback(
    [Output('main-div', 'style'),
     Output('main-div', 'className')],
    Input('theme-switch', 'value')
)
def update_theme(theme_value):
    base_style = {
        'padding': '20px',
        'minHeight': '100vh',
        'transition': 'all 0.3s ease-in-out'
    }
    
    if 'night' in theme_value:
        return {
            **base_style,
            'backgroundColor': '#121212',
            'color': '#e0e0e0',
        }, 'night-mode'
    else:
        return {
            **base_style,
            'backgroundColor': '#ffffff',
            'color': '#2d3436',
        }, 'day-mode'

@app.callback(
    Output("last-update", "children"),
    [Input("refresh-button", "n_clicks"),
     Input("days-dropdown", "value"),
     Input("period-dropdown", "value")]
)
def refresh_data(n_clicks, days, period):
    if n_clicks > 0:
        update_data(days=days, period=period)
        run_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return f"Last updated: {run_timestamp}"
    else:
        return "Data not updated yet. Please click refresh."

@app.callback(
    Output('current-price-header', 'children'),
    Input('asset-dropdown', 'value')
)
def update_current_price_header(selected_asset):
    if not global_metrics or selected_asset not in global_metrics:
        return "No data available for this asset."
    current_price = global_metrics[selected_asset].get("Current Price")
    return f"Current Price for {selected_asset}: {current_price:.6f}" if current_price is not None else "N/A"

@app.callback(
    Output('arima-rf-metrics', 'children'),
    Input('asset-dropdown', 'value')
)
def update_arima_rf_metrics(selected_asset):
    if not global_metrics or selected_asset not in global_metrics:
        return html.P("No data available for this asset.")
    predicted = global_metrics[selected_asset].get("Predicted", {})
    pred_return = predicted.get("Predicted Return")
    rmse = predicted.get("Predicted RMSE")
    r2 = predicted.get("Predicted R2")
    var_95 = predicted.get("VaR (95%)")
    current_price = global_metrics[selected_asset].get("Current Price")
    if pred_return is None or current_price is None:
        return html.P("ARIMA + RF model results not available for this asset.")
    predicted_price = current_price * np.exp(pred_return)
    return html.Div([
        html.H4("ARIMA + Random Forest Ensemble"),
        html.P(f"Predicted Next Day Price: {predicted_price:.6f}" if predicted_price is not None else "N/A"),
        html.P(f"Predicted Log Return: {pred_return:.6f}" if pred_return is not None else "N/A"),
        html.P(f"Predicted RMSE: {rmse:.6f}" if rmse is not None else "N/A"),
        html.P(f"Predicted R²: {r2:.6f}" if r2 is not None else "N/A"),
        html.P(f"Value at Risk (95%): {var_95:.6f}" if var_95 is not None else "N/A")
    ], style={'padding': '10px', 'border': '1px solid #ccc', 'marginBottom': '10px'})

@app.callback(
    Output('gbm-metrics', 'children'),
    Input('asset-dropdown', 'value')
)
def update_gbm_metrics(selected_asset):
    if selected_asset not in global_data:
        return html.P("No data available for this asset.")
    gbm_pred, gbm_rmse, gbm_r2, gbm_var = predict_gbm(selected_asset)
    current_price = global_metrics[selected_asset].get("Current Price")
    if gbm_pred is None or current_price is None:
        return html.P("GBM model results not available for this asset.")
    predicted_price = current_price * np.exp(gbm_pred)
    return html.Div([
        html.H4("Gradient Boosting Machine (GBM)"),
        html.P(f"Predicted Next Day Price: {predicted_price:.6f}"),
        html.P(f"Predicted Log Return: {gbm_pred:.6f}"),
        html.P(f"Predicted RMSE: {gbm_rmse:.6f}"),
        html.P(f"Predicted R²: {gbm_r2:.6f}"),
        html.P(f"Value at Risk (95%): {gbm_var:.6f}")
    ])

@app.callback(
    Output('additional-predicted-metrics', 'children'),
    Input('asset-dropdown', 'value')
)
def update_additional_predicted_metrics(selected_asset):
    if selected_asset not in global_data:
        return html.P("No data available for this asset.")
    lin_pred, lin_rmse, lin_r2, lin_var = predict_linear_regression(selected_asset)
    svr_pred, svr_rmse, svr_r2, svr_var = predict_svr(selected_asset)
    current_price = global_metrics[selected_asset].get("Current Price")
    lin_pred_price = current_price * np.exp(lin_pred) if lin_pred is not None and current_price is not None else None
    svr_pred_price = current_price * np.exp(svr_pred) if svr_pred is not None and current_price is not None else None
    return html.Div([
        html.H3(f"Additional Predictive Models for {selected_asset}"),
        html.Div([
            html.H4("Linear Regression"),
            html.P(f"Predicted Next Day Price: {lin_pred_price:.6f}" if lin_pred_price is not None else "N/A"),
            html.P(f"Predicted Log Return: {lin_pred:.6f}" if lin_pred is not None else "N/A"),
            html.P(f"RMSE: {lin_rmse:.6f}" if lin_rmse is not None else "N/A"),
            html.P(f"R²: {lin_r2:.6f}" if lin_r2 is not None else "N/A"),
            html.P(f"VaR (95%): {lin_var:.6f}" if lin_var is not None else "N/A"),
        ], style={'padding': '10px', 'border': '1px solid #ccc', 'marginBottom': '10px'}),
        html.Div([
            html.H4("Support Vector Regression (SVR)"),
            html.P(f"Predicted Next Day Price: {svr_pred_price:.6f}" if svr_pred_price is not None else "N/A"),
            html.P(f"Predicted Log Return: {svr_pred:.6f}" if svr_pred is not None else "N/A"),
            html.P(f"RMSE: {svr_rmse:.6f}" if svr_rmse is not None else "N/A"),
            html.P(f"R²: {svr_r2:.6f}" if svr_r2 is not None else "N/A"),
            html.P(f"VaR (95%): {svr_var:.6f}" if svr_var is not None else "N/A"),
        ], style={'padding': '10px', 'border': '1px solid #ccc', 'marginBottom': '10px'}),
    ])

def get_chart_layout(title, template, show_legend=True):
    return {
        'title': title,
        'template': template,
        'margin': dict(l=40, r=40, t=40, b=40),
        'paper_bgcolor': 'rgba(0,0,0,0)',
        'plot_bgcolor': 'rgba(0,0,0,0)',
        'showlegend': show_legend,
        'legend': dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01
        ),
        'xaxis': dict(
            showgrid=True,
            gridwidth=1,
            gridcolor='rgba(128,128,128,0.2)',
            showline=True,
            linewidth=1,
            linecolor='rgba(128,128,128,0.4)'
        ),
        'yaxis': dict(
            showgrid=True,
            gridwidth=1,
            gridcolor='rgba(128,128,128,0.2)',
            showline=True,
            linewidth=1,
            linecolor='rgba(128,128,128,0.4)'
        )
    }

@app.callback(
    Output('rolling-beta-chart', 'figure'),
    [Input("refresh-button", "n_clicks"), 
     Input('asset-dropdown', 'value'),
     Input('theme-switch', 'value')]
)
def update_rolling_beta_chart(n_clicks, selected_asset, theme_value):
    if not global_data or selected_asset not in global_data:
        return px.line(title=f"No data available for {selected_asset}")
    alt_data = global_data[selected_asset]
    btc_data = global_data["BTC"]
    alt_returns = compute_frequency_returns(alt_data.copy(), 'D')
    btc_returns = compute_frequency_returns(btc_data.copy(), 'D')
    rolling_beta_series = calculate_rolling_beta(alt_returns, btc_returns, window=30)
    if rolling_beta_series is None or rolling_beta_series.empty:
        return px.line(title=f"Rolling beta failed for {selected_asset}")
    rolling_beta_df = rolling_beta_series.reset_index()
    rolling_beta_df.columns = ['time', 'rolling_beta']
    template = 'plotly_dark' if 'night' in theme_value else 'plotly_white'
    fig = px.line(rolling_beta_df, x='time', y='rolling_beta', 
                  title=f"{selected_asset} Rolling Beta (30-day)")
    
    fig.update_layout(**get_chart_layout("Rolling Beta", template))
    return fig

@app.callback(
    Output('volatility-chart', 'figure'),
    [Input("refresh-button", "n_clicks"), 
     Input('asset-dropdown', 'value'),
     Input('theme-switch', 'value')]
)
def update_volatility_chart(n_clicks, selected_asset, theme_value):
    if not global_data or selected_asset not in global_data:
        return px.line(title=f"No data available for {selected_asset}")
    alt_data = global_data[selected_asset]
    alt_returns = compute_frequency_returns(alt_data.copy(), 'D')
    rolling_vol = alt_returns.rolling(window=30).std().dropna()
    if rolling_vol.empty:
        return px.line(title=f"Insufficient data for {selected_asset}")
    rolling_vol_df = rolling_vol.reset_index()
    rolling_vol_df.columns = ['time', 'volatility']
    template = 'plotly_dark' if 'night' in theme_value else 'plotly_white'
    fig = px.line(rolling_vol_df, x='time', y='volatility', title=f"{selected_asset} Rolling Volatility (30-day)")
    
    fig.update_layout(**get_chart_layout("Rolling Volatility", template))
    return fig

@app.callback(
    Output('rolling-alpha-chart', 'figure'),
    [Input("refresh-button", "n_clicks"), 
     Input('asset-dropdown', 'value'),
     Input('theme-switch', 'value')]
)
def update_rolling_alpha_chart(n_clicks, selected_asset, theme_value):
    if not global_data or selected_asset not in global_data:
        return px.line(title=f"No data available for {selected_asset}")
    btc_data = global_data["BTC"]
    if selected_asset == "BTC":
        btc_returns = compute_frequency_returns(btc_data.copy(), 'D')
        dates = btc_returns.index[30:] if len(btc_returns) >= 30 else []
        alpha_series = pd.Series(0, index=dates)
    else:
        alt_data = global_data[selected_asset]
        alt_returns = compute_frequency_returns(alt_data.copy(), 'D')
        btc_returns = compute_frequency_returns(btc_data.copy(), 'D')
        alpha_series = calculate_rolling_alpha(alt_returns, btc_returns, window=30)
    if alpha_series.empty:
        return px.line(title=f"Rolling alpha failed for {selected_asset}")
    rolling_alpha_df = alpha_series.reset_index()
    rolling_alpha_df.columns = ['time', 'rolling_alpha']
    template = 'plotly_dark' if 'night' in theme_value else 'plotly_white'
    fig = px.line(rolling_alpha_df, x='time', y='rolling_alpha', title=f"{selected_asset} Rolling Alpha (30-day)")
    
    fig.update_layout(**get_chart_layout("Rolling Alpha", template))
    return fig

@app.callback(
    Output('rolling-correlation-chart', 'figure'),
    [Input("refresh-button", "n_clicks"), 
     Input('asset-dropdown', 'value'),
     Input('theme-switch', 'value')]
)
def update_rolling_correlation_chart(n_clicks, selected_asset, theme_value):
    if not global_data or selected_asset not in global_data:
        return px.line(title=f"No data available for {selected_asset}")
    alt_data = global_data[selected_asset]
    btc_data = global_data["BTC"]
    alt_returns = compute_frequency_returns(alt_data.copy(), 'D')
    btc_returns = compute_frequency_returns(btc_data.copy(), 'D')
    rolling_corr_series = calculate_rolling_correlation(alt_returns, btc_returns, window=30)
    if rolling_corr_series is None or rolling_corr_series.empty:
        return px.line(title=f"Rolling correlation failed for {selected_asset}")
    rolling_corr_df = rolling_corr_series.reset_index()
    rolling_corr_df.columns = ['time', 'rolling_correlation']
    template = 'plotly_dark' if 'night' in theme_value else 'plotly_white'
    fig = px.line(rolling_corr_df, x='time', y='rolling_correlation', title=f"{selected_asset} Rolling Correlation (30-day)")
    
    fig.update_layout(**get_chart_layout("Rolling Correlation", template))
    return fig

@app.callback(
    Output('correlation-heatmap', 'figure'),
    [Input("refresh-button", "n_clicks"),
     Input('theme-switch', 'value')]
)
def update_correlation_heatmap(n_clicks, theme_value):
    if not global_data:
        return go.Figure(data=go.Heatmap(), layout=go.Layout(title="No data available for correlation heatmap"))
    daily_returns = {}
    for asset in global_data.keys():
        returns = compute_frequency_returns(global_data[asset].copy(), 'D')
        if not returns.empty:
            daily_returns[asset] = returns
    if len(daily_returns) < 2:
        return go.Figure(data=go.Heatmap(), layout=go.Layout(title="Insufficient assets for correlation heatmap"))
    df = pd.DataFrame(daily_returns)
    corr_matrix = df.corr()
    if corr_matrix.empty:
        return go.Figure(data=go.Heatmap(), layout=go.Layout(title="Correlation matrix computation failed"))
    z_values = corr_matrix.values
    x_labels = corr_matrix.columns.tolist()
    y_labels = corr_matrix.index.tolist()
    text_values = np.round(z_values, 2)
    template = 'plotly_dark' if 'night' in theme_value else 'plotly_white'
    fig = go.Figure(data=go.Heatmap(
        z=z_values,
        x=x_labels,
        y=y_labels,
        colorscale='Viridis',
        text=text_values,
        texttemplate="%{text}",
        hoverinfo="z+text"
    ))
    fig.update_layout(
        title="Correlation Heatmap Across Assets",
        xaxis_title="Asset",
        yaxis_title="Asset",
        template=template
    )
    return fig

@app.callback(
    [Output('candlestick-volume-chart', 'figure'),
     Output('technical-indicators-chart', 'figure'),
     Output('bollinger-bands-chart', 'figure')],
    [Input('asset-dropdown', 'value'),
     Input('theme-switch', 'value'),
     Input('refresh-button', 'n_clicks')]
)
def update_technical_charts(selected_asset, theme_value, n_clicks):
    if not global_data or selected_asset not in global_data:
        return [go.Figure()] * 3
    
    df = global_data[selected_asset].copy()
    df = calculate_technical_indicators(df)
    
    template = 'plotly_dark' if 'night' in theme_value else 'plotly_white'
    
    # Candlestick with Volume
    fig1 = make_subplots(rows=2, cols=1, shared_xaxes=True, 
                        vertical_spacing=0.03, 
                        row_heights=[0.7, 0.3])
    
    fig1.add_trace(go.Candlestick(
        x=df.index,
        open=df['open'],
        high=df['high'],
        low=df['low'],
        close=df['close'],
        name='OHLC'
    ), row=1, col=1)
    
    fig1.add_trace(go.Bar(
        x=df.index,
        y=df['volume'],
        name='Volume'
    ), row=2, col=1)
    
    fig1.update_layout(**get_chart_layout(f"{selected_asset} Price and Volume", template))
    
    # Technical Indicators
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=df.index, y=df['RSI'], name='RSI'))
    fig2.add_trace(go.Scatter(x=df.index, y=df['MACD'], name='MACD'))
    fig2.add_trace(go.Scatter(x=df.index, y=df['Signal_Line'], name='Signal'))
    
    fig2.update_layout(**get_chart_layout(f"{selected_asset} Technical Indicators", template))
    
    # Bollinger Bands
    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=df.index, y=df['close'], name='Price'))
    fig3.add_trace(go.Scatter(x=df.index, y=df['BB_upper'], name='Upper BB'))
    fig3.add_trace(go.Scatter(x=df.index, y=df['BB_lower'], name='Lower BB',
                             fill='tonexty'))
    
    fig3.update_layout(**get_chart_layout(f"{selected_asset} Bollinger Bands", template))
    
    return fig1, fig2, fig3

@app.callback(
    Output('risk-metrics-content', 'children'),
    [Input('asset-dropdown', 'value'),
     Input('refresh-button', 'n_clicks')]
)
def update_risk_metrics(selected_asset, n_clicks):
    if not global_data or selected_asset not in global_data:
        return html.P("No data available for risk metrics calculation.")
    
    df = global_data[selected_asset].copy()
    risk_metrics = calculate_risk_metrics(df)
    
    return html.Div([
        html.Div([
            html.H4("Risk Metrics", className='metric-header'),
            html.P(f"Maximum Drawdown: {risk_metrics['Max_Drawdown']:.2%}", className='metric-value'),
            html.P(f"Monthly Sharpe Ratio: {risk_metrics['Monthly_Sharpe']:.2f}", className='metric-value'),
            html.P(f"Sortino Ratio: {risk_metrics['Sortino_Ratio']:.2f}", className='metric-value'),
            html.P(f"Average Monthly Return: {risk_metrics['Avg_Monthly_Return']:.2%}", className='metric-value'),
            html.P(f"Monthly Volatility: {risk_metrics['Monthly_Volatility']:.2%}", className='metric-value')
        ], className='metric-section')
    ])

def calculate_risk_metrics(df):
    """
    Calculate risk metrics with detailed explanations
    """
    # Calculate daily returns
    df['daily_returns'] = df['close'].pct_change()
    returns = df['daily_returns'].dropna()
    
    # Value at Risk (VaR)
    # VaR represents the maximum loss at a confidence level
    var_95 = np.percentile(returns, 5)  # 95% confidence level
    var_99 = np.percentile(returns, 1)  # 99% confidence level
    
    # Expected Shortfall (ES) / Conditional VaR
    # Average loss beyond VaR
    es_95 = returns[returns <= var_95].mean()
    es_99 = returns[returns <= var_99].mean()
    
    # Maximum Drawdown
    # Largest peak-to-trough decline
    rolling_max = df['close'].expanding().max()
    drawdowns = df['close']/rolling_max - 1
    max_drawdown = drawdowns.min()
    
    # Monthly Returns for Sharpe Ratio
    # Resample to monthly frequency
    monthly_returns = df['close'].resample('M').last().pct_change()
    
    # Annualized Monthly Return and Volatility
    avg_monthly_return = monthly_returns.mean()
    monthly_volatility = monthly_returns.std()
    
    # Monthly Sharpe Ratio (matching Excel)
    # Using monthly risk-free rate
    monthly_rf = 0.02/12  # Assuming 2% annual risk-free rate
    monthly_sharpe = (avg_monthly_return - monthly_rf) / monthly_volatility
    
    # Sortino Ratio
    # Only considers downside volatility
    downside_returns = monthly_returns[monthly_returns < 0]
    sortino = (avg_monthly_return - monthly_rf) / downside_returns.std()
    
    # Print detailed calculations for verification
    print("\nRisk Metrics Calculation Details:")
    print(f"Average Monthly Return: {avg_monthly_return:.4f}")
    print(f"Monthly Volatility: {monthly_volatility:.4f}")
    print(f"Monthly Risk-Free Rate: {monthly_rf:.4f}")
    print(f"Monthly Sharpe Ratio: {monthly_sharpe:.4f}")
    
    return {
        'VaR_95': var_95,
        'VaR_99': var_99,
        'ES_95': es_95,
        'ES_99': es_99,
        'Max_Drawdown': max_drawdown,
        'Monthly_Sharpe': monthly_sharpe,
        'Sortino_Ratio': sortino,
        'Avg_Monthly_Return': avg_monthly_return,
        'Monthly_Volatility': monthly_volatility
    }

def calculate_technical_indicators(df):
    df = df.copy()
    df['time'] = pd.to_datetime(df['time_period_start'])
    df.set_index('time', inplace=True)
    
    # Convert price columns to numeric
    price_cols = ['price_open', 'price_high', 'price_low', 'price_close']
    for col in price_cols:
        df[col] = pd.to_numeric(df[col])
    
    # Rename columns for convenience
    df = df.rename(columns={
        'price_open': 'open',
        'price_high': 'high',
        'price_low': 'low',
        'price_close': 'close',
        'volume_traded': 'volume'
    })
    
    # Calculate RSI
    delta = df['close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
    rs = gain / loss
    df['RSI'] = 100 - (100 / (1 + rs))
    
    # Calculate MACD
    exp1 = df['close'].ewm(span=12, adjust=False).mean()
    exp2 = df['close'].ewm(span=26, adjust=False).mean()
    df['MACD'] = exp1 - exp2
    df['Signal_Line'] = df['MACD'].ewm(span=9, adjust=False).mean()
    
    # Calculate Bollinger Bands
    df['BB_middle'] = df['close'].rolling(window=20).mean()
    bb_std = df['close'].rolling(window=20).std()
    df['BB_upper'] = df['BB_middle'] + (bb_std * 2)
    df['BB_lower'] = df['BB_middle'] - (bb_std * 2)
    
    return df

# Main Execution
if __name__ == "__main__":
    logging.info("Waiting for manual refresh to update data...")
    logging.info("Dash is running on http://127.0.0.1:8050/")
    print("Waiting for manual refresh to update data...")
    print("Dash is running on http://127.0.0.1:8050/")
    app.run_server(debug=True)

